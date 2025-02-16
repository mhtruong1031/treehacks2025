import openpyxl, requests, os, zipfile

from bs4 import BeautifulSoup

# Configurables
UNFIILTERED_GENETIC_DATA_PATH = 'genetic_data_unfiltered.xlsx'
GENETIC_DATA_DIR_PATH         = 'resources/genetic_data'
MUTUAL_IDS_PATH               = 'mutual_ids.txt'

def main():
    # Intialize workbook
    with open(MUTUAL_IDS_PATH, 'r') as f:
        mutual_ids = f.read().splitlines()

    wb = openpyxl.load_workbook(UNFIILTERED_GENETIC_DATA_PATH)
    gd = wb["Genetic Data (unfiltered)"]

    # Scrape links
    hyperlinks = []
    for i in range(len(gd['A'])):
        if gd.cell(i+1, 1).value in mutual_ids:
            hyperlinks.append(gd.cell(i+1, 6).hyperlink.target)

    # Take links and save data into dir
    for url in hyperlinks:
        r    = requests.get(url=url)
        if r.status_code != 401:
            soup = BeautifulSoup(r.content)
            cmd  = soup.find(id = 'wget-example').contents

            os.system(cmd[0][2:] + f" -P {GENETIC_DATA_DIR_PATH}")

    # Clear zip files
    for file in os.listdir(GENETIC_DATA_DIR_PATH):
        if 'zip' in file:
            with zipfile.ZipFile(f'{GENETIC_DATA_DIR_PATH}/{file}', 'r') as zip_ref:
                zip_ref.extractall(GENETIC_DATA_DIR_PATH)
            os.system(f'rm {GENETIC_DATA_DIR_PATH}/{file}')

if __name__ == '__main__':
    main()

    