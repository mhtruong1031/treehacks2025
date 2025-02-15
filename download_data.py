import openpyxl

with open("mutual_ids.txt", "r") as f:
    mutual_ids = f.read().splitlines()

print(mutual_ids)
wb = openpyxl.load_workbook('genetic_data_unfiltered.xlsx')
gd = wb["Genetic Data (unfiltered)"]

for i in range(len(gd['A'])):
    if gd.cell(i+1, 1).value in mutual_ids:
        print(gd.cell(i+1, 6).hyperlink)